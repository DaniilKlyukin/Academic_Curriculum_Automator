import os
import re
import logging
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional, Union, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from docx import Document

logger = logging.getLogger(__name__)


class ExcelConfig:
    """Конфигурация столбцов и имен листов Excel по умолчанию."""
    COMPETENCY_SHEET_NAME: str = "Компетенции"
    PLAN_SHEET_NAME: str = "ПланСвод"  # Приоритетный лист по умолчанию

    # Дефолтная колонка для текста содержания (если не найдена автопоиском)
    COMP_TXT_COL: int = 5

    PLAN_IDX_COL: int = 2
    PLAN_NAME_COL: int = 3


class CompetencyReportGenerator:
    """Класс для сопоставления компетенций и дисциплин из учебного плана в документ Word."""

    def __init__(self, excel_path: Union[str, Path], word_path: Union[str, Path]) -> None:
        self.excel_path: Path = Path(excel_path)
        self.word_path: Path = Path(word_path)

    def _find_semester_columns(self, sheet: Worksheet) -> Dict[int, int]:
        """
        Сканирует первые строки листа для автоматического определения столбцов семестров (18-25).
        """
        semester_cols: Dict[int, int] = {}
        for col in range(1, sheet.max_column + 1):
            for row in range(1, 6):
                cell_val: str = str(sheet.cell(row=row, column=col).value or "").lower()
                if "сем" in cell_val:
                    match = re.search(r"сем.*?(\d+)", cell_val)
                    if match:
                        sem_num: int = int(match.group(1))
                        semester_cols[sem_num] = col
                        break
        return semester_cols

    def _find_assessment_columns(self, sheet: Worksheet) -> List[int]:
        """
        Находит индексы колонок промежуточной аттестации (Экзамен, Зачет, Зачет с оц., КР, Реферат).
        """
        cols: List[int] = []
        for col in range(1, sheet.max_column + 1):
            for row in range(1, 6):
                val: str = str(sheet.cell(row=row, column=col).value or "").lower()
                if any(x in val for x in ["экзамен", "зачет", "кр", "реферат", "кп"]):
                    cols.append(col)
                    break
        return list(set(cols))

    def _find_competency_text_column(self, sheet: Worksheet) -> int:
        """
        Ищет столбец 'Содержание' на листе компетенций.
        """
        txt_col: int = ExcelConfig.COMP_TXT_COL
        for col in range(1, sheet.max_column + 1):
            for row in range(1, 11):
                val: str = str(sheet.cell(row=row, column=col).value or "").lower().strip()
                if "содержание" in val:
                    return col
        return txt_col

    @staticmethod
    def _extract_semesters_from_assessment(val: Any) -> List[int]:
        """
        Извлекает номера семестров из ячеек промежуточной аттестации (например, '123' -> [1, 2, 3]).
        """
        if val is None:
            return []
        val_str: str = str(val).strip()
        semesters: List[int] = []

        if val_str.isdigit():
            for char in val_str:
                sem_num = int(char)
                if 1 <= sem_num <= 12:
                    semesters.append(sem_num)
        else:
            parts = re.split(r'[,\s;]+', val_str)
            for part in parts:
                if part.isdigit():
                    sem_num = int(part)
                    if 1 <= sem_num <= 12:
                        semesters.append(sem_num)
        return list(set(semesters))

    def _parse_plan_sheet(self, sheet: Worksheet) -> Dict[str, Dict[str, Any]]:
        """
        Парсит учебный план, собирая названия предметов и семестры их проведения.
        """
        plan_db: Dict[str, Dict[str, Any]] = {}
        semester_cols: Dict[int, int] = self._find_semester_columns(sheet)
        assessment_cols: List[int] = self._find_assessment_columns(sheet)

        logger.info(f"Найдены столбцы семестров: {list(semester_cols.keys())}")
        logger.info(f"Найдены столбцы аттестации (индексы): {assessment_cols}")

        for row in range(1, sheet.max_row + 1):
            idx_val: str = str(sheet.cell(row=row, column=ExcelConfig.PLAN_IDX_COL).value or "").strip()
            if re.match(r"^(Б\d|ФТД)", idx_val):
                name_val: str = str(sheet.cell(row=row, column=ExcelConfig.PLAN_NAME_COL).value or "").strip()

                active_semesters: List[int] = []

                # 1. Считываем семестры из колонок нагрузки
                for sem_num, col_idx in semester_cols.items():
                    val: Optional[Any] = sheet.cell(row=row, column=col_idx).value
                    if val is not None:
                        try:
                            if float(val) > 0:
                                active_semesters.append(sem_num)
                        except ValueError:
                            pass

                # 2. Дополнительно считываем семестры из колонок аттестации
                for col_idx in assessment_cols:
                    val_assess: Optional[Any] = sheet.cell(row=row, column=col_idx).value
                    if val_assess is not None:
                        active_semesters.extend(self._extract_semesters_from_assessment(val_assess))

                plan_db[idx_val] = {
                    "name": name_val,
                    "semesters": sorted(list(set(active_semesters)))
                }
        return plan_db

    def _parse_competencies(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Парсит иерархическую структуру компетенций, сканируя три разные колонки структуры.
        """
        competencies: List[Dict[str, Any]] = []
        current_comp: Optional[Dict[str, Any]] = None

        txt_col = self._find_competency_text_column(sheet)
        logger.info(f"Определен столбец Содержания компетенций -> {txt_col}")

        comp_pattern = re.compile(r"^([A-Za-zА-Яа-я]+)-\d+$")
        indicator_pattern = re.compile(r"^([A-Za-zА-Яа-я]+)-\d+\.\d+$")

        print("\n" + "=" * 50)
        print("ДЕТАЛЬНЫЙ ЛОГ ИМПОРТА ИЗЛИСТА 'КОМПЕТЕНЦИИ':")
        print("=" * 50)

        for row in range(1, sheet.max_row + 1):
            # Извлекаем значения из трех потенциальных колонок иерархии индексов
            val_col1: str = str(sheet.cell(row=row, column=1).value or "").strip()
            val_col2: str = str(sheet.cell(row=row, column=2).value or "").strip()
            val_col3: str = str(sheet.cell(row=row, column=3).value or "").strip()
            text_val: str = str(sheet.cell(row=row, column=txt_col).value or "").strip()

            # Сценарий 1: Компетенция находится в первой колонке (A)
            if val_col1 and comp_pattern.match(val_col1):
                current_comp = {
                    "comp_code": val_col1,
                    "comp_name": text_val,
                    "indicators": [],
                    "mapped_codes": set()
                }
                competencies.append(current_comp)
                print(f"Строка {row:03} | [Компетенция] {val_col1} -> {text_val[:50]}...")

            # Сценарий 2: Индикатор находится во второй колонке (B)
            elif val_col2 and indicator_pattern.match(val_col2):
                if current_comp:
                    current_comp["indicators"].append(f"{val_col2} {text_val}")
                    print(f"Строка {row:03} |   [Индикатор]   {val_col2} -> {text_val[:50]}...")
                else:
                    print(f"Строка {row:03} |   [ВНИМАНИЕ] Пропущен индикатор без компетенции: {val_col2}")

            # Сценарий 3: Дисциплина/Практика находится в третьей колонке (C)
            elif val_col3 and re.match(r"^(Б\d|ФТД)", val_col3):
                if current_comp:
                    current_comp["mapped_codes"].add(val_col3)
                    print(f"Строка {row:03} |     [Привязка]    {val_col3} -> {text_val[:50]}...")
                else:
                    print(f"Строка {row:03} |     [ВНИМАНИЕ] Пропущена дисциплина без компетенции: {val_col3}")

        print("=" * 50 + "\n")
        return competencies

    @staticmethod
    def _format_semesters(sem_list: List[int]) -> str:
        """Форматирует список семестров в строку."""
        if not sem_list:
            return ""
        sorted_sems: List[int] = sorted(list(set(sem_list)))
        sems_str: str = ", ".join(map(str, sorted_sems))
        if len(sorted_sems) == 1:
            return f"{sems_str} семестр"
        return f"{sems_str} семестры"

    def generate(self) -> None:
        """
        Запускает процесс обработки и сохранения результатов.
        """
        if not self.excel_path.exists():
            print(f"Ошибка: Исходный файл {self.excel_path.name} не найден по указанному пути.")
            return

        try:
            wb = load_workbook(str(self.excel_path.absolute()), data_only=True)
        except Exception as e:
            logger.error(f"Не удалось открыть файл Excel: {e}")
            return

        # Интеллектуальный поиск листа учебного плана
        plan_sheet_name: str = ExcelConfig.PLAN_SHEET_NAME
        if plan_sheet_name not in wb.sheetnames:
            if "План" in wb.sheetnames:
                plan_sheet_name = "План"
            else:
                for name in wb.sheetnames:
                    if "план" in name.lower():
                        plan_sheet_name = name
                        break

        if ExcelConfig.COMPETENCY_SHEET_NAME not in wb.sheetnames or plan_sheet_name not in wb.sheetnames:
            logger.error(f"В книге отсутствуют необходимые листы. Доступные листы: {wb.sheetnames}")
            return

        print(f"Анализ структуры учебного плана на листе '{plan_sheet_name}'...")
        plan_db: Dict[str, Dict[str, Any]] = self._parse_plan_sheet(wb[plan_sheet_name])

        sheet_comp = wb[ExcelConfig.COMPETENCY_SHEET_NAME]
        competencies: List[Dict[str, Any]] = self._parse_competencies(sheet_comp)

        if not competencies:
            print("\n[ВНИМАНИЕ] Данные по компетенциям не были импортированы.")
            return

        total: int = len(competencies)
        print(f"\n{'№':<9} | {'Статус':<8} | {'Код':<8} | {'Компетенция'}")
        print("-" * 100)

        # Инициализация Word
        doc = Document()

        # Настройка альбомной ориентации для широкой таблицы
        section = doc.sections[-1]
        new_width, new_height = section.page_height, section.page_width
        section.page_width = new_width
        section.page_height = new_height

        table = doc.add_table(rows=1, cols=5)
        table.style = "Table Grid"

        # Заголовки таблицы Word
        hdr_cells = table.rows[0].cells
        hdr_cells[0].text = "Код и наименование компетенции"
        hdr_cells[1].text = "Код и наименование индикатора достижения компетенции"
        hdr_cells[2].text = "Дисциплины (модули)"
        hdr_cells[3].text = "Практики"
        hdr_cells[4].text = "Семестр формирования"

        for i, comp in enumerate(competencies, 1):
            comp_code: str = comp["comp_code"]
            status: str = "OK"

            try:
                row_cells = table.add_row().cells

                # 1. Запись компетенции
                row_cells[0].text = f"{comp_code} {comp['comp_name']}"

                # 2. Запись индикаторов
                row_cells[1].text = "\n".join(comp["indicators"])

                # Использование списков с последующим исключением дубликатов по названию предмета
                disciplines_names: List[str] = []
                practices_names: List[str] = []
                disc_semesters: Set[int] = set()
                prac_semesters: Set[int] = set()

                for item_code in sorted(list(comp["mapped_codes"])):
                    item_info: Optional[Dict[str, Any]] = plan_db.get(item_code)
                    if not item_info:
                        name: str = f"{item_code} (название не найдено)"
                        sems: List[int] = []
                    else:
                        name = item_info["name"]
                        sems = item_info["semesters"]

                    # Распределение на дисциплины и практики
                    if item_code.startswith("Б1") or item_code.startswith("ФТД"):
                        if name not in disciplines_names:
                            disciplines_names.append(name)
                        disc_semesters.update(sems)
                    elif item_code.startswith("Б2") or item_code.startswith("Б3"):
                        if name not in practices_names:
                            practices_names.append(name)
                        prac_semesters.update(sems)

                # 3. Дисциплины (упорядоченно и без повторов)
                row_cells[2].text = "\n".join(sorted(disciplines_names))

                # 4. Практики (упорядоченно и без повторов)
                row_cells[3].text = "\n".join(sorted(practices_names))

                # 5. Семестры
                sem_info: List[str] = []
                if disc_semesters:
                    sem_info.append(f"Дисциплины — {self._format_semesters(list(disc_semesters))}")
                if prac_semesters:
                    sem_info.append(f"Практики — {self._format_semesters(list(prac_semesters))}")

                row_cells[4].text = "\n".join(sem_info)

            except Exception as e:
                status = "ERR"
                logger.error(f"Непредвиденная ошибка при обработке {comp_code}: {e}")

            comp_desc: str = comp['comp_name'][:50]
            print(f"[{i:03}/{total:03}] | {status:<8} | {comp_code:<8} | {comp_desc}...")

        try:
            self.word_path.parent.mkdir(parents=True, exist_ok=True)
            doc.save(str(self.word_path.absolute()))
            print(f"\nПроцесс завершен. Таблица успешно сохранена в: {self.word_path.name}")
        except Exception as e:
            logger.error(f"Не удалось сохранить итоговый документ Word: {e}")


if __name__ == "__main__":
    # Настройка логирования по умолчанию
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    print("=== Панель управления генерацией отчетов ===")

    # Интерактивный ввод путей
    user_excel_path: str = input("Шаг 1. Введите путь к исходному файлу Excel (например, plan.xlsx): ").strip()
    user_word_path: str = input("Шаг 2. Введите путь для сохранения файла Word (например, result.docx): ").strip()

    # Использование значений по умолчанию при пустом вводе
    if not user_excel_path:
        user_excel_path = "plan.xlsx"
        print(f"Используется путь по умолчанию для Excel: {user_excel_path}")

    if not user_word_path:
        user_word_path = "competencies_table.docx"
        print(f"Используется путь по умолчанию для Word: {user_word_path}")

    print("\nЗапуск процесса генерации...")
    generator = CompetencyReportGenerator(
        excel_path=user_excel_path,
        word_path=user_word_path
    )
    generator.generate()