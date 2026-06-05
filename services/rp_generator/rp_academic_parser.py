import os
import re
import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl import load_workbook

# Настройка логирования
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)


def clean_text(val: Any) -> str:
    """Очищает строку от технических артефактов Excel XML (_x000D_ и др.) и лишних пробелов."""
    if val is None:
        return ""
    s = str(val).strip()
    # Удаление артефактов кодирования Excel XML типа _x000D_, _x000D_ и схожих
    s = re.sub(r"_x[0-9a-fA-F]{4}_", "", s)
    # Замена повторяющихся пробелов и переносов на один стандартный пробел
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def parse_semesters_string(sem_str: str) -> List[int]:
    """Интеллектуально разбирает строку семестров контроля (например, '12345', '78', '1-5', '1, 2')."""
    if not sem_str:
        return []

    sem_str = sem_str.strip()
    semesters = set()

    # 1. Если строка состоит только из цифр без разделителей
    if sem_str.isdigit():
        for char in sem_str:
            val = int(char)
            if 1 <= val <= 12:
                semesters.add(val)
        return sorted(list(semesters))

    # 2. Обработка диапазонов вида "1-5"
    range_match = re.match(r"^(\d+)\s*[-—–]\s*(\d+)$", sem_str)
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        if start <= end and end <= 12:
            return list(range(start, end + 1))

    # 3. Разделение по разделителям
    parts = re.split(r'[,\s;]+', sem_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue

        if part.isdigit():
            if len(part) > 1 and int(part) > 12:
                for char in part:
                    val = int(char)
                    if 1 <= val <= 12:
                        semesters.add(val)
            else:
                val = int(part)
                if 1 <= val <= 12:
                    semesters.add(val)

    return sorted(list(semesters))


class AcademicPlanParser:
    """Парсер учебной нагрузки, метаданных титульного листа и структуры плана из Excel."""

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path

    def _find_plan_sheet(self, wb) -> str:
        """Ищет подходящий лист детального учебного плана, строго исключая сводные листы."""
        # 1. Сначала ищем строгое совпадение с "План"
        if "План" in wb.sheetnames:
            return "План"

        # 2. Ищем лист, содержащий слово "план", но НЕ содержащий слово "свод"
        for name in wb.sheetnames:
            name_lower = name.lower()
            if "план" in name_lower and "свод" not in name_lower:
                return name

        # 3. Резервный поиск любого листа со словом "план"
        for name in wb.sheetnames:
            if "план" in name.lower():
                return name

        raise ValueError(f"Детальный лист учебного плана ('План') не найден в книге. Доступные листы: {wb.sheetnames}")

    def _parse_department_directory(self, wb) -> Dict[int, str]:
        """Строит справочник 'Код кафедры -> Название кафедры' на основе листа 'ПланСвод'."""
        dept_map = {}
        if "ПланСвод" not in wb.sheetnames:
            return dept_map

        sheet = wb["ПланСвод"]
        col_code = 26  # Дефолтная колонка Кода в ПланСвод
        col_name = 27  # Дефолтная колонка Наименования в ПланСвод

        # Динамический поиск колонок Кода и Наименования на случай смещения
        for col in range(1, sheet.max_column + 1):
            for row in range(1, 4):
                val = clean_text(sheet.cell(row=row, column=col).value).lower()
                if "закреплен" in val or "кафедр" in val:
                    sub_val = clean_text(sheet.cell(row=3, column=col).value).lower()
                    if "код" in sub_val:
                        col_code = col
                    elif "наимен" in sub_val:
                        col_name = col

        # Наполнение справочника
        for row in range(4, sheet.max_row + 1):
            code_val = sheet.cell(row=row, column=col_code).value
            name_val = sheet.cell(row=row, column=col_name).value
            try:
                code_int = int(float(code_val)) if code_val is not None else None
            except (ValueError, TypeError):
                code_int = None

            if code_int is not None and name_val:
                dept_map[code_int] = clean_text(name_val)

        logger.info(f"Справочник кафедр успешно построен из 'ПланСвод' (всего записей: {len(dept_map)})")
        return dept_map

    def _map_semester_columns(self, sheet) -> Dict[int, Dict[str, int]]:
        """Автоматически определяет номера колонок нагрузки (Лек, Лаб, Пр, СР) для каждого семестра."""
        semester_map = {}
        for col in range(1, sheet.max_column + 1):
            cell_val = clean_text(sheet.cell(row=2, column=col).value)
            if "семестр" in cell_val.lower():
                m = re.search(r"семестр\s*(\d+)", cell_val, re.IGNORECASE)
                if m:
                    sem_num = int(m.group(1))
                    semester_map[sem_num] = {
                        "Lek": col,
                        "Lab": col + 1,
                        "Pr": col + 2,
                        "CP": col + 3
                    }
        return semester_map

    def _find_department_column(self, sheet) -> int:
        """Автоматически находит индекс колонки кода закрепленной кафедры."""
        for col in range(1, sheet.max_column + 1):
            for row in range(1, 4):
                cell_val = clean_text(sheet.cell(row=row, column=col).value).lower()
                if "закреплен" in cell_val:
                    return col
        return 50

    def _extract_field_by_keyword(self, sheet, regex: re.Pattern, max_cols_offset: int = 15) -> str:
        """Помехоустойчиво ищет значение параметра на листе по ключевому регулярному выражению."""
        for row in range(1, min(sheet.max_row + 1, 150)):  # Ограничение сканирования титула первыми 150 строками
            for col in range(1, sheet.max_column + 1):
                cell_val = clean_text(sheet.cell(row=row, column=col).value)
                if not cell_val:
                    continue
                if regex.search(cell_val):
                    # Если значение указано в этой же ячейке после двоеточия
                    if ":" in cell_val:
                        parts = cell_val.split(":", 1)
                        val = clean_text(parts[1])
                        if val and len(val) > 1:
                            return val
                    # Иначе ищем значение в ячейках справа
                    for offset in range(1, max_cols_offset):
                        if col + offset <= sheet.max_column:
                            test_val = clean_text(sheet.cell(row=row, column=col + offset).value)
                            if test_val and len(test_val) > 1:
                                return test_val
                    # Резерв: ищем значение в ячейке прямо под текущей (на случай вертикального расположения)
                    if row + 1 <= sheet.max_row:
                        test_val = clean_text(sheet.cell(row=row + 1, column=col).value)
                        if test_val and len(test_val) > 1:
                            return test_val
        return ""

    def _extract_title_metadata(self, wb) -> Dict[str, str]:
        """Парсит лист 'Титул' и собирает метаданные учебного плана для генерации заголовков РП."""
        metadata = {
            "qualification": "",
            "education_form": "",
            "study_duration": "",
            "start_year": "",
            "academic_year": "",
            "fgos_standard": "",
            "faculty": "",
            "department": "",
            "direction_code": "",
            "direction_name": "",
            "profile": ""
        }

        if "Титул" not in wb.sheetnames:
            logger.warning("Лист 'Титул' отсутствует в книге. Пропуск сбора общих метаданных.")
            return metadata

        sheet = wb["Титул"]

        # Паттерны поиска метаданных
        pat_qualification = re.compile(r"квалификация", re.IGNORECASE)
        pat_education_form = re.compile(r"форма\s+обучения", re.IGNORECASE)
        pat_study_duration = re.compile(r"срок\s+(?:получения|обучения)", re.IGNORECASE)
        pat_start_year = re.compile(r"год\s+начала\s+подготовки", re.IGNORECASE)
        pat_academic_year = re.compile(r"учебный\s+год", re.IGNORECASE)
        pat_fgos = re.compile(r"(?:стандарт|фгос)\b", re.IGNORECASE)
        pat_faculty = re.compile(r"факультет\b|институт\b", re.IGNORECASE)
        pat_department = re.compile(r"кафедра\b", re.IGNORECASE)
        pat_profile = re.compile(r"профиль\b|направленность\b", re.IGNORECASE)
        pat_direction = re.compile(r"направление\b|специальность\b|код\s+направления\b|шифр\s+направления\b",
                                   re.IGNORECASE)

        # Сбор данных с листа "Титул"
        metadata["qualification"] = self._extract_field_by_keyword(sheet, pat_qualification)
        metadata["education_form"] = self._extract_field_by_keyword(sheet, pat_education_form)
        metadata["study_duration"] = self._extract_field_by_keyword(sheet, pat_study_duration)
        metadata["start_year"] = self._extract_field_by_keyword(sheet, pat_start_year)
        metadata["academic_year"] = self._extract_field_by_keyword(sheet, pat_academic_year)
        metadata["fgos_standard"] = self._extract_field_by_keyword(sheet, pat_fgos)
        metadata["faculty"] = self._extract_field_by_keyword(sheet, pat_faculty)
        metadata["department"] = self._extract_field_by_keyword(sheet, pat_department)
        metadata["profile"] = self._extract_field_by_keyword(sheet, pat_profile)

        # Специализированный поиск направления подготовки (Код + Наименование)
        direction_code = ""
        direction_name = ""
        found_code_row = None

        # 1. Поиск кода направления вида "XX.XX.XX" в первых 60 строках листа
        for r in range(1, min(sheet.max_row + 1, 60)):
            for c in range(1, sheet.max_column + 1):
                val = clean_text(sheet.cell(row=r, column=c).value)
                if val:
                    m = re.match(r"^\s*(\d{2}\.\d{2}\.\d{2})\s*$", val)
                    if m:
                        direction_code = m.group(1)
                        found_code_row = r
                        break
            if found_code_row:
                break

        # 2. Поиск названия направления на строки ниже найденного кода
        if found_code_row:
            # Просматриваем до 3 строк ниже строки с кодом направления
            for offset in range(1, 4):
                target_row = found_code_row + offset
                if target_row > sheet.max_row:
                    break

                # Собираем все непустые ячейки в этой строке
                row_vals = []
                for col in range(1, sheet.max_column + 1):
                    cell_val = clean_text(sheet.cell(row=target_row, column=col).value)
                    if cell_val:
                        row_vals.append(cell_val)

                # Исключаем строки, содержащие другие служебные метаданные (например, Профиль, Кафедра и др.)
                valid_vals = []
                for val in row_vals:
                    val_lower = val.lower()
                    if any(kw in val_lower for kw in [
                        "профиль", "направленность", "кафедра", "факультет",
                        "квалификация", "утверждаю", "форма обучения", "протокол"
                    ]):
                        continue
                    if len(val) > 2:
                        valid_vals.append(val)

                if valid_vals:
                    direction_name = valid_vals[0]
                    break

        # Резервный вариант: если точный код не найден по шаблону, используем поиск по ключевому слову
        if not direction_code and not direction_name:
            direction_raw = self._extract_field_by_keyword(sheet, pat_direction)
            if direction_raw:
                code_match = re.search(r"(\d{2}\.\d{2}\.\d{2})", direction_raw)
                if code_match:
                    direction_code = code_match.group(1)
                    name_part = direction_raw.replace(direction_code, "").strip()
                    direction_name = re.sub(r"^[-\s,.:\)]+", "", name_part).strip()
                else:
                    direction_name = direction_raw

        metadata["direction_code"] = direction_code
        metadata["direction_name"] = clean_text(direction_name)

        logger.info(f"Успешно извлечены метаданные плана: {metadata}")
        return metadata

    def parse(self) -> Dict[str, Any]:
        """Парсит лист плана и возвращает структуру академической нагрузки и метаданные блоков."""
        wb = load_workbook(str(self.excel_path.absolute()), data_only=True)

        # Сбор общих метаданных учебного плана (форма обучения, квалификация, стандарты и др.)
        metadata = self._extract_title_metadata(wb)

        # Строим справочник соответствия кодов и названий кафедр из ПланСвод
        dept_directory = self._parse_department_directory(wb)

        sheet_name = self._find_plan_sheet(wb)
        sheet = wb[sheet_name]
        logger.info(f"Анализ учебной нагрузки на листе: '{sheet_name}'")

        # Базовые колонки (структура ИжГТУ)
        col_index = 2  # Код дисциплины (например, Б1.О.11)
        col_name = 3  # Наименование дисциплины

        # Колонки форм контроля
        col_exam = 4
        col_credit = 5
        col_graded_credit = 6
        col_kp = 7
        col_kr = 8

        # === ДИНАМИЧЕСКИЙ ПОИСК КОЛОНОК (Автоопределение сдвигов и скрытых столбцов) ===
        col_plan_hours = None
        col_total_srs = None
        col_total_control = None
        col_credits = None

        for col in range(1, sheet.max_column + 1):
            for r in range(1, 6):
                val_raw = sheet.cell(row=r, column=col).value
                val = clean_text(val_raw).lower() if val_raw else ""
                if not val:
                    continue

                # 1. Поиск ЗЕТ ("Экспертное" в группе "з.е.")
                if "эксперт" in val:
                    parent_val_1 = clean_text(sheet.cell(row=1, column=col).value).lower()
                    parent_val_2 = clean_text(sheet.cell(row=2, column=col).value).lower()
                    if any(x in parent_val_1 or x in parent_val_2 for x in ["з.е.", "зачетн"]):
                        col_credits = col

                # 2. Поиск "По плану"
                elif "по плану" in val:
                    col_plan_hours = col

                # 3. Поиск общей СРС "СР" (в группе акад. часов)
                elif val == "ср" or "самостоятельная" in val:
                    parent_val_1 = clean_text(sheet.cell(row=1, column=col).value).lower()
                    parent_val_2 = clean_text(sheet.cell(row=2, column=col).value).lower()
                    if any(x in parent_val_1 or x in parent_val_2 for x in ["итого", "акад"]):
                        col_total_srs = col

                # 4. Поиск "Контроль" или "Конт роль" (в группе акад. часов)
                elif any(x in val for x in ["контроль", "конт роль", "конт.роль"]):
                    parent_val_1 = clean_text(sheet.cell(row=1, column=col).value).lower()
                    parent_val_2 = clean_text(sheet.cell(row=2, column=col).value).lower()
                    if any(x in parent_val_1 or x in parent_val_2 for x in ["итого", "акад"]):
                        col_total_control = col

        # Резервные дефолты (на случай несовпадения формата)
        if col_credits is None: col_credits = 10
        if col_plan_hours is None: col_plan_hours = 13
        if col_total_srs is None: col_total_srs = 16
        if col_total_control is None: col_total_control = 17

        logger.info(
            f"Определены индексы колонок: з.е.={col_credits}, По плану={col_plan_hours}, СР={col_total_srs}, Контроль={col_total_control}")

        col_dept = self._find_department_column(sheet)
        logger.info(f"Определен столбец кода закрепленной кафедры -> {col_dept}")

        semester_cols = self._map_semester_columns(sheet)
        logger.info(f"Обнаружена структура семестровых колок: {list(semester_cols.keys())}")

        disciplines = {}

        # Переменные для отслеживания текущего контекста иерархии
        current_block = ""
        current_part = ""
        current_elective_group: Optional[Dict[str, str]] = None

        for row in range(4, sheet.max_row + 1):
            idx_val = clean_text(sheet.cell(row=row, column=col_index).value)

            # --- СЦЕНАРИЙ А: Строка заголовка (Блок или Подблок) ---
            if not idx_val:
                row_texts = [clean_text(sheet.cell(row=row, column=col).value) for col in range(1, 6)]
                full_row_text = " ".join([t for t in row_texts if t])

                # 1. Поиск названия Блока
                if re.search(r"Блок\s+\d", full_row_text, re.IGNORECASE) or re.search(r"^ФТД\.", full_row_text,
                                                                                      re.IGNORECASE):
                    for t in row_texts:
                        if re.search(r"Блок\s+\d", t, re.IGNORECASE) or re.search(r"^ФТД\.", t, re.IGNORECASE):
                            current_block = t
                            current_part = ""  # Сбрасываем часть при смене блока
                            current_elective_group = None
                            logger.info(f"Текущий Блок -> {current_block}")
                            break
                    continue

                # 2. Поиск названия Части (Подблока)
                if "обязательная часть" in full_row_text.lower():
                    current_part = "Обязательная часть"
                    current_elective_group = None
                    logger.info(f"Текущая Часть -> {current_part}")
                    continue
                elif "часть, формируемая" in full_row_text.lower():
                    current_part = "Часть, формируемая участниками образовательных отношений"
                    current_elective_group = None
                    logger.info(f"Текущая Часть -> {current_part}")
                    continue

                continue

            # --- СЦЕНАРИЙ Б: Строка элемента (Дисциплина, Практика, ГИА, Факультатив) ---
            if re.match(r"^(Б\d|ФТД)", idx_val):
                name_val = clean_text(sheet.cell(row=row, column=col_name).value)
                if not name_val:
                    continue

                # 1. Определение заголовка элективной группы (Дисциплина по выбору)
                if re.match(r"^Б\d+\.[ОВ]\.ДВ\.\d+$", idx_val):
                    current_elective_group = {
                        "code": idx_val,
                        "name": name_val
                    }
                    logger.info(f"Обнаружена группа выбора -> {idx_val}: {name_val}")
                    continue  # Пропускаем строку-заголовок (она не несет нагрузки)

                # Проверка: относится ли текущая строка к активной группе элективов
                active_elective = None
                if current_elective_group:
                    if idx_val.startswith(current_elective_group["code"]):
                        active_elective = current_elective_group
                    else:
                        current_elective_group = None  # Вышли за пределы группы

                dept_val = sheet.cell(row=row, column=col_dept).value
                try:
                    department_code = int(float(dept_val)) if dept_val is not None else None
                except (ValueError, TypeError):
                    department_code = None

                if department_code is None:
                    continue

                # Извлечение текстового названия кафедры по ее коду через справочник ПланСвод
                department_name = dept_directory.get(department_code, "")

                # 3. Извлечение зачетных единиц (з.е.) из 9-го столбца
                credits_cell = sheet.cell(row=row, column=col_credits).value
                try:
                    credit_units = float(credits_cell) if credits_cell is not None else 0.0
                    if credit_units.is_integer():
                        credit_units = int(credit_units)
                except (ValueError, TypeError):
                    credit_units = 0

                # 4. Извлечение форм контроля
                exam_sems = clean_text(sheet.cell(row=row, column=col_exam).value)
                credit_sems = clean_text(sheet.cell(row=row, column=col_credit).value)
                graded_credit_sems = clean_text(sheet.cell(row=row, column=col_graded_credit).value)
                kp_sems = clean_text(sheet.cell(row=row, column=col_kp).value)
                kr_sems = clean_text(sheet.cell(row=row, column=col_kr).value)

                # 5. Распределение семестровой нагрузки
                load_by_semester = {}
                total_lectures = 0
                total_labs = 0
                total_practicals = 0
                total_cp = 0
                total_control_hours = 0
                total_control_self_study = 0.0

                # Списки семестров контроля
                exam_sems = parse_semesters_string(exam_sems)
                credit_sems = parse_semesters_string(credit_sems)
                graded_credit_sems = parse_semesters_string(graded_credit_sems)
                kp_sems = parse_semesters_string(kp_sems)
                kr_sems = parse_semesters_string(kr_sems)

                # Функция для округления до 1 знака после запятой без сохранения .0
                def format_academic_hours(value: float) -> float | int:
                    rounded = round(value, 1)
                    return int(rounded) if rounded.is_integer() else rounded

                for sem_num, cols in semester_cols.items():
                    lek = sheet.cell(row=row, column=cols["Lek"]).value
                    lab = sheet.cell(row=row, column=cols["Lab"]).value
                    pr = sheet.cell(row=row, column=cols["Pr"]).value
                    cp = sheet.cell(row=row, column=cols["CP"]).value

                    try:
                        lek_h = int(float(lek)) if lek else 0
                        lab_h = int(float(lab)) if lab else 0
                        pr_h = int(float(pr)) if pr else 0
                        cp_h = float(cp) if cp else 0.0  # Считываем СРС как float (сохраняем .25)
                    except (ValueError, TypeError):
                        continue

                    # Определение базовой формы промежуточного контроля
                    control_info = None
                    if sem_num in exam_sems:
                        control_info = {
                            "type": "Exam",
                            "kcha": 0.4,
                            "self_study": 35.6,
                            "total": 36.0
                        }
                    elif sem_num in graded_credit_sems:
                        control_info = {
                            "type": "GradedCredit",
                            "kcha": 0.4,
                            "self_study": 1.6,
                            "total": 2.0
                        }
                    elif sem_num in credit_sems:
                        control_info = {
                            "type": "Credit",
                            "kcha": 0.3,
                            "self_study": 1.7,
                            "total": 2.0
                        }

                    # Учет дополнительных часов промежуточного контроля для Курсовых проектов (КП) и Курсовых работ (КР)
                    cw_kp_kcha = 0.0
                    if sem_num in kp_sems:
                        cw_kp_kcha += 0.75
                    if sem_num in kr_sems:
                        cw_kp_kcha += 0.75

                    if cw_kp_kcha > 0:
                        if control_info is None:
                            control_info = {
                                "type": "CourseWorkProject",
                                "kcha": cw_kp_kcha,
                                "self_study": 0.0,
                                "total": cw_kp_kcha
                            }
                        else:
                            control_info["kcha"] = round(control_info["kcha"] + cw_kp_kcha, 2)
                            control_info["total"] = round(control_info["total"] + cw_kp_kcha, 2)

                    # Семестр активен, если есть аудиторная нагрузка, СРС или форма контроля
                    if lek_h > 0 or lab_h > 0 or pr_h > 0 or cp_h > 0 or control_info is not None:
                        load_by_semester[str(sem_num)] = {
                            "lectures": lek_h,
                            "laboratory_works": lab_h,
                            "practical_classes": pr_h,
                            "self_study": format_academic_hours(cp_h),
                            "intermediate_control": control_info
                        }
                        total_lectures += lek_h
                        total_labs += lab_h
                        total_practicals += pr_h
                        total_cp += cp_h

                        if control_info:
                            total_control_hours += control_info["total"]
                            total_control_self_study += control_info["self_study"]

                # Прямое считывание эталонных сумм из таблицы Excel во избежание накопительных погрешностей
                plan_hours_cell = sheet.cell(row=row, column=col_plan_hours).value
                total_srs_cell = sheet.cell(row=row, column=col_total_srs).value

                try:
                    direct_plan_hours = int(float(plan_hours_cell)) if plan_hours_cell is not None else 0
                except (ValueError, TypeError):
                    direct_plan_hours = 0

                try:
                    direct_total_srs = float(total_srs_cell) if total_srs_cell is not None else 0.0
                except (ValueError, TypeError):
                    direct_total_srs = 0.0

                # Приоритет отдаем эталонным значениям из строк плана Excel
                final_self_study = (direct_total_srs if direct_total_srs > 0 else total_cp) + total_control_self_study
                final_total = direct_plan_hours if direct_plan_hours > 0 else (
                            total_lectures + total_labs + total_practicals + total_cp + total_control_hours)

                final_self_study = format_academic_hours(final_self_study)
                final_total = format_academic_hours(final_total)

                disciplines[idx_val] = {
                    "code": idx_val,
                    "name": name_val,
                    "department_code": department_code,
                    "department_name": department_name,
                    "credit_units": credit_units,
                    "structure": {
                        "block": current_block,
                        "part": current_part,
                        "elective_group": active_elective
                    },
                    "control_forms": {
                        "exams": exam_sems,
                        "credits": credit_sems,
                        "graded_credits": graded_credit_sems,
                        "course_projects": kp_sems,
                        "course_works": kr_sems
                    },
                    "total_hours": {
                        "lectures": total_lectures,
                        "laboratory_works": total_labs,
                        "practical_classes": total_practicals,
                        "self_study": final_self_study,
                        "total": final_total
                    },
                    "load_by_semester": load_by_semester
                }

        return {
            "metadata": metadata,
            "disciplines": disciplines
        }


def main():
    print("=== Модуль разбора академической нагрузки и планов ===")
    user_excel = input("Введите путь к файлу Excel (например, plan.xlsx): ").strip()
    if not user_excel:
        user_excel = "plan.xlsx"

    excel_path = Path(user_excel)
    if not excel_path.exists():
        print("Ошибка: Файл плана не найден.")
        return

    user_output_dir = input(
        "Введите путь к папке для сохранения результатов (по умолчанию 'services/rp_generator'): ").strip()
    if not user_output_dir:
        user_output_dir = "services/rp_generator"

    parser = AcademicPlanParser(excel_path)
    try:
        data = parser.parse()

        output_dir = Path(user_output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "rp_academic_workload.json"

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"[Успешно] Данные по академической нагрузке сохранены в:\n{output_path.absolute()}")
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}", exc_info=True)


if __name__ == "__main__":
    main()