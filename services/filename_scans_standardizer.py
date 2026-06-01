import os
import re
import logging
import io
import difflib
from pathlib import Path
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from PIL import Image
import fitz
import pytesseract

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(TESSERACT_CMD):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
else:
    logger.warning(f"Указанный путь к Tesseract OCR не найден: {TESSERACT_CMD}")


def normalize_homoglyphs(text: str) -> str:
    """Заменяет английские буквы-омоглифы на их русские аналоги."""
    eng_to_rus = {
        'a': 'а', 'b': 'в', 'c': 'с', 'e': 'е', 'h': 'н', 'k': 'к', 'm': 'м',
        'o': 'о', 'p': 'р', 't': 'т', 'x': 'х', 'y': 'у', 'ё': 'е'
    }
    return "".join(eng_to_rus.get(char, char) for char in text.lower())


def clean_text_strict(text: str) -> str:
    """Жесткая очистка текста для алгоритма скользящего окна."""
    if not text:
        return ""
    s = text.lower()
    s = normalize_homoglyphs(s)
    s = re.sub(r'[^а-яa-z0-9\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def abbreviate_word(word: str) -> str:
    """Сокращает слово до первой согласной после первой гласной с поддержкой спецсимволов."""
    vowels = set("аеёиоуыэюя")
    word_clean = "".join([c for c in word if c.isalnum() or c in ('+', '#')])
    if not word_clean:
        return ""

    first_vowel_idx = -1
    for idx, char in enumerate(word_clean):
        if char.lower() in vowels:
            first_vowel_idx = idx
            break

    if first_vowel_idx == -1:
        return word_clean

    first_consonant_after_vowel_idx = -1
    for idx in range(first_vowel_idx + 1, len(word_clean)):
        if word_clean[idx].lower() not in vowels:
            first_consonant_after_vowel_idx = idx
            break

    if first_consonant_after_vowel_idx == -1:
        return word_clean

    return word_clean[:first_consonant_after_vowel_idx + 1]


def abbreviate_discipline(name: str) -> str:
    """Сокращает название дисциплины (идентично для сканов и DOCX)."""
    normalized_name = name.replace("-", " ")
    words = normalized_name.split()

    stop_words = {"на", "по", "в", "с", "для", "под", "о", "об", "за", "из", "от", "до", "без"}

    abbr_words = []
    for w in words:
        if w.lower() in stop_words:
            continue

        abbr = abbreviate_word(w)
        if not abbr:
            continue

        if w.lower() == "и":
            abbr_words.append("и")
        else:
            abbr_words.append(abbr.capitalize())

    return "".join(abbr_words)


def load_disciplines_from_excel(excel_path: Path) -> List[str]:
    wb = load_workbook(str(excel_path.absolute()), data_only=True)
    disciplines = []

    target_sheets = [name for name in wb.sheetnames if "план" in name.lower()]

    for sheet_name in target_sheets:
        sheet = wb[sheet_name]
        for row in range(1, sheet.max_row + 1):
            code_val = str(sheet.cell(row=row, column=2).value or "").strip()
            name_val = str(sheet.cell(row=row, column=3).value or "").strip()
            if code_val and name_val:
                if re.match(r"^[БФТД]\d+", code_val):
                    if name_val not in disciplines:
                        disciplines.append(name_val)
    return list(set(disciplines))


def extract_text_from_file(file_path: Path) -> str:
    ext = file_path.suffix.lower()
    text = ""

    try:
        if ext == ".pdf":
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text() or ""

            if not text.strip() and len(doc) > 0:
                page = doc[0]
                pix = page.get_pixmap(dpi=150)
                img_bytes = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_bytes))
                text = pytesseract.image_to_string(img, lang="rus")

        elif ext in (".jpg", ".jpeg", ".png"):
            img = Image.open(file_path).convert('RGB')
            text = pytesseract.image_to_string(img, lang="rus")

    except Exception as e:
        logger.warning(f"Ошибка распознавания файла {file_path.name}: {e}")

    return text


def find_best_match_sliding_window(ocr_text: str, plan_disciplines: List[str]) -> Tuple[Optional[str], float]:
    cleaned_ocr = clean_text_strict(ocr_text)
    ocr_words = cleaned_ocr.split()

    if not ocr_words:
        return None, 0.0

    if "итогов" in cleaned_ocr and "аттестац" in cleaned_ocr or \
       "выпускн" in cleaned_ocr and "квалификацион" in cleaned_ocr:
        for disc in plan_disciplines:
            dl = clean_text_strict(disc)
            if any(kw in dl for kw in ["выпускн", "квалификацион", "вкр", "защит", "аттестац", "гиа"]):
                return disc, 1.0

    best_match = None
    max_ratio = 0.0

    for disc in plan_disciplines:
        cleaned_disc = clean_text_strict(disc)
        disc_words = cleaned_disc.split()
        n_words = len(disc_words)

        if n_words == 0:
            continue

        window_sizes = set([max(1, n_words - 1), n_words, n_words + 1])
        disc_max_ratio = 0.0

        for w_size in window_sizes:
            for i in range(len(ocr_words) - w_size + 1):
                window_text = " ".join(ocr_words[i:i + w_size])
                ratio = difflib.SequenceMatcher(None, cleaned_disc, window_text).ratio()

                if ratio > disc_max_ratio:
                    disc_max_ratio = ratio

        if disc_max_ratio > max_ratio:
            max_ratio = disc_max_ratio
            best_match = disc

    return best_match, max_ratio


class ScanRenamer:

    def __init__(self, scans_dir: str, excel_path: str):
        self.scans_dir = Path(scans_dir)
        self.excel_path = Path(excel_path)

    def run(self):
        if not self.scans_dir.exists() or not self.excel_path.exists():
            logger.error("Проверьте правильность путей к папке со сканами и файлу Excel.")
            return

        logger.info("Загрузка списка дисциплин из Excel...")
        plan_disciplines = load_disciplines_from_excel(self.excel_path)
        logger.info(f"Загружено дисциплин: {len(plan_disciplines)}")

        valid_extensions = {".jpg", ".jpeg", ".png", ".pdf"}
        files = sorted([
            f for f in self.scans_dir.iterdir()
            if f.suffix.lower() in valid_extensions and not f.name.startswith("~$")
        ])

        triples = [files[i:i + 3] for i in range(0, len(files), 3)]

        logger.info("\n=== ЭТАП 1: Анализ текста сканов ===")
        group_results = []

        for idx, group in enumerate(triples, 1):
            if len(group) < 3:
                continue

            file_names = ", ".join([f.name for f in group])
            logger.info(f"\n--- Анализ группы сканов №{idx} ({file_names}) ---")

            best_group_match = None
            highest_confidence = 0.0

            for file_path in group:
                logger.info(f"  Читаем файл: {file_path.name}...")
                text = extract_text_from_file(file_path)

                matched_disc, confidence = find_best_match_sliding_window(text, plan_disciplines)

                if matched_disc and confidence > highest_confidence:
                    highest_confidence = confidence
                    best_group_match = matched_disc

                if highest_confidence >= 0.90:
                    logger.info(f"  [+] Найдено уверенное совпадение: '{best_group_match}' ({highest_confidence:.1%})")
                    break

            group_results.append({
                'idx': idx,
                'files': group,
                'match': best_group_match if highest_confidence >= 0.70 else None,
                'conf': highest_confidence
            })

        logger.info("\n=== ЭТАП 2: Разрешение конфликтов и дубликатов ===")
        claims = {}
        for res in group_results:
            if res['match']:
                claims.setdefault(res['match'], []).append(res)

        for disc, claiming_groups in claims.items():
            if len(claiming_groups) > 1:
                claiming_groups.sort(key=lambda x: x['conf'], reverse=True)
                winner = claiming_groups[0]

                logger.warning(f"[!] Конфликт для '{disc}': претендуют {len(claiming_groups)} группы.")
                logger.info(f"    -> Победитель: Группа №{winner['idx']} (уверенность {winner['conf']:.1%})")

                for loser in claiming_groups[1:]:
                    logger.warning(f"    -> Сброс группы №{loser['idx']} (уверенность была {loser['conf']:.1%}) до 'Неизвестная'.")
                    loser['match'] = None
                    loser['conf'] = 0.0

        logger.info("\n=== ЭТАП 3: Переименование файлов ===")
        updated_count = 0

        for res in group_results:
            idx = res['idx']
            group = res['files']
            match = res['match']

            if match:
                current_discipline_abbr = abbreviate_discipline(match)
                logger.info(f"  [+] Группа №{idx} утверждена: '{match}' -> префикс ({current_discipline_abbr})")
            else:
                current_discipline_abbr = f"НеизвестнаяДисциплинаГруппа{idx}"
                logger.warning(f"  [-] Группа №{idx} не распознана. Присвоено тех. имя: {current_discipline_abbr}")

            for page_num, file_path in enumerate(group, 1):
                new_name = f"{current_discipline_abbr}{page_num}{file_path.suffix}"
                new_path = file_path.with_name(new_name)

                counter = 1
                while new_path.exists():
                    stem = Path(new_name).stem
                    new_path = file_path.with_name(f"{stem}_{counter}{file_path.suffix}")
                    counter += 1

                try:
                    file_path.rename(new_path)
                    logger.info(f"      Переименован: '{file_path.name}' -> '{new_path.name}'")
                    updated_count += 1
                except Exception as e:
                    logger.error(f"      [!] Ошибка переименования '{file_path.name}': {e}")

        print(f"\n[Готово] Обработка сканов завершена. Успешно переименовано файлов: {updated_count}")


def main():
    print("=== Интеллектуальное переименование сканов РП ===")
    try:
        user_scans_dir = input("Шаг 1. Введите путь к папке со сканами РП (картинки/pdf): ").strip().strip('"')
        user_excel_path = input("Шаг 2. Введите путь к файлу учебного плана Excel (plan.xlsx): ").strip().strip('"')
    except UnicodeDecodeError:
        print("Ошибка кодировки консоли! Используются пути по умолчанию (.) и (plan.xlsx).")
        user_scans_dir = "."
        user_excel_path = "plan.xlsx"

    if not user_scans_dir:
        user_scans_dir = "."
    if not user_excel_path:
        user_excel_path = "plan.xlsx"

    renamer = ScanRenamer(user_scans_dir, user_excel_path)
    renamer.run()


if __name__ == "__main__":
    main()