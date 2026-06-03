import os
import re
import json
import logging
from pathlib import Path
from typing import Any, Dict, Set, Tuple

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)


def _find_competency_text_column(sheet) -> int:
    """Ищет столбец 'Содержание' на листе компетенций."""
    txt_col = 5  # Дефолтная колонка по умолчанию
    for col in range(1, sheet.max_column + 1):
        for row in range(1, 11):
            val = str(sheet.cell(row=row, column=col).value or "").lower().strip()
            if "содержание" in val:
                return col
    return txt_col


class CompetencyMapper:
    """Парсер листа 'Компетенции' для сопоставления дисциплин и индикаторов.

    Разделяет данные на реестр описаний и реляционную карту связей.
    """

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path

    def parse(self) -> Dict[str, Any]:
        wb = load_workbook(str(self.excel_path.absolute()), data_only=True)
        if "Компетенции" not in wb.sheetnames:
            raise ValueError(f"Лист 'Компетенции' отсутствует в книге. Доступные листы: {wb.sheetnames}")

        sheet = wb["Компетенции"]
        txt_col = _find_competency_text_column(sheet)
        logger.info(f"Определен столбец Содержания компетенций -> {txt_col}")

        # Помехоустойчивые паттерны к разным типам тире/дефисов и пробелов
        comp_pattern = re.compile(r"^([A-Za-zА-Яа-я]+)\s*[-–—]\s*(\d+)$")
        indicator_pattern = re.compile(r"^([A-Za-zА-Яа-я]+)\s*[-–—]\s*(\d+\.\d+)$")
        item_pattern = re.compile(r"^(Б\d|ФТД)")

        # Внутренние структуры для сбора данных
        competencies_registry: Dict[str, Dict[str, Any]] = {}
        subject_to_competencies: Dict[str, Dict[str, Any]] = {}

        current_comp_code = ""
        current_comp_text = ""
        current_ind_code = ""
        current_ind_text = ""

        max_scan_rows = max(sheet.max_row, 1000)
        consecutive_empty_rows = 0

        for row in range(1, max_scan_rows + 1):
            # Адаптивное считывание первых 4-х колонок из-за смещения структуры ПК/ОПК
            val_col1 = str(sheet.cell(row=row, column=1).value or "").strip()
            val_col2 = str(sheet.cell(row=row, column=2).value or "").strip()
            val_col3 = str(sheet.cell(row=row, column=3).value or "").strip()
            val_col4 = str(sheet.cell(row=row, column=4).value or "").strip()
            text_val = str(sheet.cell(row=row, column=txt_col).value or "").strip()

            if not val_col1 and not val_col2 and not val_col3 and not val_col4:
                consecutive_empty_rows += 1
                if consecutive_empty_rows > 50:
                    break
                continue

            consecutive_empty_rows = 0

            # 1. Поиск КОМПЕТЕНЦИИ (может быть в колонке 1 или 2)
            comp_code = None
            if val_col1 and comp_pattern.match(val_col1):
                comp_code = val_col1
            elif val_col2 and comp_pattern.match(val_col2):
                comp_code = val_col2

            if comp_code:
                # Нормализация кода (замена дефисов на стандартный дефис)
                current_comp_code = re.sub(r'[-–—]', '-', comp_code).replace(" ", "")
                current_comp_text = text_val
                current_ind_code = ""
                current_ind_text = ""

                if current_comp_code not in competencies_registry:
                    competencies_registry[current_comp_code] = {
                        "competency_code": current_comp_code,
                        "competency_text": current_comp_text,
                        "disciplines": set(),
                        "indicators": {}
                    }
                continue

            # 2. Поиск ИНДИКАТОРА (может быть в колонке 2 или 3)
            indicator_code = None
            if val_col2 and indicator_pattern.match(val_col2):
                indicator_code = val_col2
            elif val_col3 and indicator_pattern.match(val_col3):
                indicator_code = val_col3

            if indicator_code:
                current_ind_code = re.sub(r'[-–—]', '-', indicator_code).replace(" ", "")
                current_ind_text = text_val

                if current_comp_code:
                    if current_comp_code not in competencies_registry:
                        competencies_registry[current_comp_code] = {
                            "competency_code": current_comp_code,
                            "competency_text": "",
                            "disciplines": set(),
                            "indicators": {}
                        }

                    indicators_dict = competencies_registry[current_comp_code]["indicators"]
                    if current_ind_code not in indicators_dict:
                        indicators_dict[current_ind_code] = {
                            "indicator_code": current_ind_code,
                            "indicator_text": current_ind_text,
                            "disciplines": set()
                        }
                continue

            # 3. Поиск ДИСЦИПЛИНЫ/ПРАКТИКИ (может быть в колонке 3 или 4)
            item_code = None
            if val_col3 and item_pattern.match(val_col3):
                item_code = val_col3
            elif val_col4 and item_pattern.match(val_col4):
                item_code = val_col4

            if item_code:
                if not current_comp_code or not current_ind_code:
                    continue

                discipline_code = item_code
                discipline_name = text_val

                # 3.1. Заполнение связей в блоке subject_to_competencies
                if discipline_code not in subject_to_competencies:
                    subject_to_competencies[discipline_code] = {
                        "discipline_code": discipline_code,
                        "discipline_name": discipline_name,
                        "competencies": {}
                    }

                sub_comp_dict = subject_to_competencies[discipline_code]["competencies"]
                if current_comp_code not in sub_comp_dict:
                    sub_comp_dict[current_comp_code] = []

                if current_ind_code not in sub_comp_dict[current_comp_code]:
                    sub_comp_dict[current_comp_code].append(current_ind_code)

                # 3.2. Заполнение обратных ссылок в реестре competencies_registry
                if current_comp_code not in competencies_registry:
                    competencies_registry[current_comp_code] = {
                        "competency_code": current_comp_code,
                        "competency_text": "",
                        "disciplines": set(),
                        "indicators": {}
                    }
                competencies_registry[current_comp_code]["disciplines"].add(discipline_code)

                indicators_dict = competencies_registry[current_comp_code]["indicators"]
                if current_ind_code not in indicators_dict:
                    indicators_dict[current_ind_code] = {
                        "indicator_code": current_ind_code,
                        "indicator_text": "",
                        "disciplines": set()
                    }
                indicators_dict[current_ind_code]["disciplines"].add(discipline_code)

        # Преобразование set во множествах в sorted list для валидного сохранения JSON
        serialized_registry = {}
        for comp_code, comp_data in competencies_registry.items():
            serialized_registry[comp_code] = {
                "competency_code": comp_data["competency_code"],
                "competency_text": comp_data["competency_text"],
                "disciplines": sorted(list(comp_data["disciplines"])),
                "indicators": {}
            }
            for ind_code, ind_data in comp_data["indicators"].items():
                serialized_registry[comp_code]["indicators"][ind_code] = {
                    "indicator_code": ind_data["indicator_code"],
                    "indicator_text": ind_data["indicator_text"],
                    "disciplines": sorted(list(ind_data["disciplines"]))
                }

        return {
            "competencies_registry": serialized_registry,
            "subject_to_competencies": subject_to_competencies
        }


def main():
    print("=== Модуль разбора карты компетенций ===")
    user_excel = input("Введите путь к файлу Excel (например, plan.xlsx): ").strip()
    if not user_excel:
        user_excel = "plan.xlsx"

    excel_path = Path(user_excel)
    if not excel_path.exists():
        print("Ошибка: Файл плана не найден.")
        return

    # Запрос пути для сохранения результатов
    user_output_dir = input(
        "Введите путь к папке для сохранения результатов (по умолчанию 'services/rp_generator'): ").strip()
    if not user_output_dir:
        user_output_dir = "services/rp_generator"

    mapper = CompetencyMapper(excel_path)
    try:
        data = mapper.parse()

        # Создание директории и формирование финального пути
        output_dir = Path(user_output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "rp_subject_competency_map.json"

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"[Успешно] Карта компетенций предметов сохранена в:\n{output_path.absolute()}")
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}", exc_info=True)


if __name__ == "__main__":
    main()